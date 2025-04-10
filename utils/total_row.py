from openpyxl.styles import Font


def add_total_row(ws, df):
    # Clean numeric-looking strings like "$12.34", "€1,000.50", "15%"
    for col in df.columns:
        if df[col].dtype == "object":
            try:
                # Remove $ € , % and convert to float
                cleaned = df[col].replace(r"[\$,€,%]", "", regex=True).replace(",", "", regex=True).astype(float)
                df[col] = cleaned
            except (ValueError, TypeError):
                continue  # If it fails, just skip (probably not numeric)

    numeric_columns = df.select_dtypes(include=["number"]).columns

    if not numeric_columns.any():
        return  # No numeric columns found, skip total row

    start_row = 4
    total_row = len(df) + start_row

    for col_name in numeric_columns:
        col_index = df.columns.get_loc(col_name) + 1
        col_letter = chr(64 + col_index)

        # Insert SUM formula
        ws[f"{col_letter}{total_row}"] = f"=SUM({col_letter}{start_row}:{col_letter}{total_row - 1})"
        ws[f"{col_letter}{total_row}"].font = Font(bold=True)

        # Label cell if possible
        if col_index > 1:
            label_cell = ws.cell(row=total_row, column=col_index - 1)
            if label_cell.value is None:
                label_cell.value = "Total:"
                label_cell.font = Font(bold=True)
