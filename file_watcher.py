import time
import json
import shutil
import pandas as pd
from datetime import datetime
from pathlib import Path
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

#  Load config
with open("config/settings.json") as config_file:
    config = json.load(config_file)

WATCH_FOLDER = Path(config["watch_folder"])
OUTPUT_FOLDER = Path(config["output_folder"])
ARCHIVE_FOLDER = Path(config["archive_folder"])

# Logger
def log(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {message}")

# Check for CSV
def is_valid_file(file_path):
    return file_path.suffix.lower() == ".csv"

# Clean columns
def clean_numeric_columns(df):
    for col in df.select_dtypes(include="object").columns:
        if df[col].astype(str).str.contains(r"^\s*[\$\€\£]?\d", na=False).any():
            df[col] = df[col].astype(str).replace(r"[^\d.\-]", "", regex=True)
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df

# Process CSV file
def process_file(file_path):
    log(f"New file detected: {file_path.name}")
    try:
        df = pd.read_csv(file_path)
        df = clean_numeric_columns(df)
        log(f"CSV file successfully read ({len(df)} rows)")

        from helpers.column_selector import get_numeric_columns
        ask_user = config.get("ask_user_for_totals", False)
        numeric_columns = get_numeric_columns(df, ask_user=ask_user)
        log(f" All numeric columns: {list(df.select_dtypes(include=['number']).columns)}")
        log(f" Columns selected for total row: {numeric_columns}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Add date
        ws["A1"] = "Report Date:"
        ws["B1"] = '=TEXT(TODAY(),"yyyy-mm-dd")'
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws.append([])

        # Write data
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        for cell in ws[3]:
            cell.font = Font(bold=True)

        # Create Excel table
        start_row = 3
        end_row = start_row + len(df)
        end_col_letter = chr(65 + len(df.columns) - 1)
        table_range = f"A{start_row}:{end_col_letter}{end_row}"

        table = Table(displayName="DataTable", ref=table_range)
        table.totalsRowShown = False  # we do it manually
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)




        #  Add manual total row (SUM formulas)

        total_row = end_row + 1
        for col_name in numeric_columns:
            col_idx = df.columns.get_loc(col_name)
            col_letter = chr(65 + col_idx)
            ws[f"A{total_row}"] = "Total:"  # Etiket
            ws[f"{col_letter}{total_row}"] = f"=SUM({col_letter}{start_row + 1}:{col_letter}{end_row})"
            ws[f"{col_letter}{total_row}"].font = Font(bold=True)

            # Add 'Total:' label to previous column if empty
            label_cell = ws.cell(row=total_row, column=col_idx + 1)
            if not label_cell.value:
                label_cell.value = "Total:"
                label_cell.font = Font(bold=True)


        # Auto column width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Save
        report_name = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = OUTPUT_FOLDER / report_name
        wb.save(report_path)
        log(f" Report generated: {report_path.name}")

        # Archive
        archive_path = ARCHIVE_FOLDER / file_path.name
        shutil.move(str(file_path), archive_path)
        log(f" Original file archived: {archive_path.name}")

    except PermissionError:
        log(" File is open in another program. Please close it.")
    except Exception as e:
        log(" ERROR while processing file:")
        log(f"   File: {file_path.name}")
        log(f"   Reason: {type(e).__name__} – {e}")

# File Watcher
class FileHandler(FileSystemEventHandler):
    def on_created(self, event):
        file_path = Path(event.src_path)
        if is_valid_file(file_path):
            time.sleep(1)
            process_file(file_path)

# Start watching
if __name__ == "__main__":
    log(" File Automation System started...")

    for folder in [WATCH_FOLDER, OUTPUT_FOLDER, ARCHIVE_FOLDER]:
        folder.mkdir(exist_ok=True)

    observer = Observer()
    observer.schedule(FileHandler(), str(WATCH_FOLDER), recursive=False)
    observer.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()
